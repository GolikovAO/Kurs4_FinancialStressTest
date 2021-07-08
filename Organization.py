import requests
import json
import zipfile
import fnmatch
import os
import openpyxl

class Organization(object):
    def __init__(self, INN):
        self.INN = INN
        self.id = self.GetIdOfOrganization()
        self.excel_filename = self.DownloadExcel()
        self.Data2019 = []
        self.Data2020 = []
        self.Balance2019 = []
        self.Balance2020 = []
        self.Name = ''
        self.Address = ''

    def GetIdOfOrganization(self):
        request_url = 'https://bo.nalog.ru/nbo/organizations/search?query=%09' + self.INN + '&page=0'
        headers = {
            'Host': 'bo.nalog.ru',
            'User-Agent': 'PostmanRuntime/7.26.8',
            'Accept': '*/*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'X-Requested-With': 'XMLHttpRequest'
        }
        request = requests.get(request_url, headers=headers)
        parsing_request = json.loads(request.text)
        id_org = parsing_request['content'][0]['id']
        return id_org

    def DownloadExcel(self):
        if not os.path.isdir(self.INN):
            url = 'https://bo.nalog.ru/download/bfo/' + str(
                self.id) + '?auditReport=true&balance=true&capitalChange=true&clarification=true&targetedFundsUsing=true&correctionNumber=0&financialResult=true&fundsMovement=true&type=XLS&period=2020'
            headers = {
                'Host': 'bo.nalog.ru',
                'User-Agent': 'PostmanRuntime/7.26.8',
                'Accept': '*/*',
                'Accept-Encoding': 'gzip, deflate, br',
                'Connection': 'keep-alive',
                'X-Requested-With': 'XMLHttpRequest'
            }
            r = requests.get(url, headers=headers)
            os.mkdir(self.INN)
            with open(str(self.INN) + "/Data.zip", "wb") as code:
                code.write(r.content)
            z = zipfile.ZipFile(self.INN + "/Data.zip", 'r')
            filepattern = '*.xlsx'
            for file in z.infolist():
                if fnmatch.fnmatch(file.filename, filepattern):
                    excel_filename = file.filename
                    z.extract(file.filename, path=self.INN)
            z.close()
            return excel_filename
        else:
            example_dir = os.path.abspath(os.curdir)
            content = os.listdir(example_dir + '/' + self.INN)
            dirs = []
            for file in content:
                if file.endswith('.xlsx'):
                    dirs.append(file)
            excel_filename = dirs[0]
            return excel_filename

    def Get2019DataFromExcel(self):
        wb = openpyxl.load_workbook(self.INN + "/" + self.excel_filename)
        sheet = wb['Financial Result']
        data = []
        for row in sheet['A7':'BA27']:
            tmp_list = []
            for cell in row:
                if (cell.value != None):
                    tmp_list.append(cell.value)
            if (tmp_list):
                data.append(tmp_list[3])
        self.Data2019 = data
        return data

    def Get2020DataFromExcel(self):
        wb = openpyxl.load_workbook(self.INN + "/" + self.excel_filename)
        sheet = wb['Financial Result']
        data = []
        for row in sheet['A7':'BA27']:
            tmp_list = []
            for cell in row:
                if (cell.value != None):
                    tmp_list.append(cell.value)
            if (tmp_list):
                data.append(tmp_list[2])
        self.Data2020 = data
        return data

    def GetBalance2019(self):
        wb = openpyxl.load_workbook(self.INN + "/" + self.excel_filename)
        sheet = wb['Balance']
        data = []
        for row in sheet['A3':'BA60']:
            tmp_list = []
            for cell in row:
                if (cell.value != None):
                    tmp_list.append(cell.value)
            if (len(tmp_list) == 5):
                tmp_list = tmp_list[0:4]
                tmp_list.pop(1)
                tmp_list.pop(1)
                data.append(tmp_list)
            elif (len(tmp_list) == 4):
                tmp_list.pop(1)
                tmp_list = tmp_list[0:3]
                data.append(tmp_list)
        self.Balance2019 = data
        return data

    def GetBalance2020(self):
        wb = openpyxl.load_workbook(self.INN + "/" + self.excel_filename)
        sheet = wb['Balance']
        data = []
        for row in sheet['A3':'BA60']:
            tmp_list = []
            for cell in row:
                if (cell.value != None):
                    tmp_list.append(cell.value)
            if (len(tmp_list) == 5):
                tmp_list = tmp_list[0:3]
                tmp_list.pop(1)
                data.append(tmp_list)
            elif (len(tmp_list) == 4):
                tmp_list = tmp_list[0:2]
                data.append(tmp_list)
        self.Balance2020 = data
        return data

    def GetNameAndAddress(self):
        wb = openpyxl.load_workbook(self.INN + "/" + self.excel_filename)
        sheet = wb['Organization Info']
        data = []
        for row in sheet['A6':'BA6']:
            tmp_list = []
            for cell in row:
                if (cell.value != None):
                    tmp_list.append(cell.value)
            if (tmp_list):
                data.append(tmp_list[1])
        self.Name = data[0]
        data = []
        for row in sheet['A16':'BA16']:
            tmp_list = []
            for cell in row:
                if (cell.value != None):
                    tmp_list.append(cell.value)
            if (tmp_list):
                data.append(tmp_list[1])
        self.Address = data[0]